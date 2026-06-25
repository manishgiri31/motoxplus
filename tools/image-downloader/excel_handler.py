from pathlib import Path
from openpyxl import load_workbook

_REQUIRED = {"Product Name", "Brand", "SKU", "Vehicle"}


class ExcelHandler:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.path}")
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self._headers: dict[str, int] = {}
        self._parse_headers()

    # ── Public ─────────────────────────────────────────────────────────────────

    def read_products(self) -> list[dict]:
        products = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            sku = row[self._col("SKU") - 1]
            if not sku:
                continue
            products.append({
                "sku":     str(sku).strip(),
                "name":    str(row[self._col("Product Name") - 1] or "").strip(),
                "brand":   str(row[self._col("Brand") - 1] or "").strip(),
                "vehicle": str(row[self._col("Vehicle") - 1] or "").strip(),
            })
        return products

    def update_paths(self, paths: dict[str, str]) -> None:
        """Write downloaded image paths into an 'Image Path' column and save."""
        if "Image Path" not in self._headers:
            col = max(self._headers.values()) + 1
            self.ws.cell(row=1, column=col, value="Image Path")
            self._headers["Image Path"] = col

        path_col = self._col("Image Path")
        sku_col  = self._col("SKU")

        for row in self.ws.iter_rows(min_row=2):
            raw = row[sku_col - 1].value
            if raw and str(raw).strip() in paths:
                row[path_col - 1].value = paths[str(raw).strip()]

        self.wb.save(self.path)

    # ── Private ────────────────────────────────────────────────────────────────

    def _parse_headers(self) -> None:
        for idx, cell in enumerate(self.ws[1], start=1):
            if cell.value:
                self._headers[str(cell.value).strip()] = idx
        missing = _REQUIRED - self._headers.keys()
        if missing:
            raise ValueError(
                f"Excel file is missing required columns: {', '.join(sorted(missing))}"
            )

    def _col(self, name: str) -> int:
        return self._headers[name]
